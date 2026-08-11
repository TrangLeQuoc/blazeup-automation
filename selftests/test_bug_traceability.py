"""The `be_gap` chain: code marker ↔ TEST_CASES docs ↔ Bug_Tracker must agree.

A `be_gap` TC is one that fails on purpose because the product has an open defect, and
it is EXCLUDED from the merge gate — it goes red and the pipeline still passes. That
makes the marker the single most dangerous line in a test file: nothing in a run tells
you it is there. The only way to notice a stale one is to compare three places by hand,
which is exactly what does not happen.

Two failures this repo actually had:

* **A closed bug kept its marker for 3 weeks.** The markers did not name a bug id, so
  when BE closed the defect there was no link to follow — the TC stayed outside the gate
  long after it would have passed. `test_the_bug_is_still_open` is that check.
* **Three TCs were `be_gap` in code while the docs never said so** (PARTNER_PORTAL_SHELL
  _001/_002, SA_PARTNER_MODULE_011 — measured 2026-08-11). Anyone reading the docs would
  believe those TCs guarded the gate. `test_doc_says_be_gap` is that check.

Adding a module needs NO change here — see "Module discovery" below.

Everything reads files only — no staging, no browser, no secrets.
"""

import re
from pathlib import Path

import openpyxl
import pytest

PROJECT_ROOT = Path(__file__).resolve().parents[1]
TESTS_DIR = PROJECT_ROOT / "tests"

# One domain today (`blazeup`, per docs/{domain}/ in utils/sync_registry.py). If a second
# domain ever appears, this constant and TRACKER become the loop — the module dimension
# below is already dynamic, the domain dimension is not.
DOCS_DIR = PROJECT_ROOT / "docs" / "blazeup"
TRACKER = DOCS_DIR / "Bug_Tracker.xlsx"

# Bug ids share ONE sequence across all modules, in ONE Bug_Tracker.xlsx: the sheet's
# "Test Case Name" column already says which module a bug belongs to, so encoding the
# module into the id as well would duplicate it. A new module keeps using BUG-API-NNN /
# BUG-UI-NNN and this pattern needs no edit.
BUG_ID_RE = re.compile(r"BUG-(?:API|UI)-\d+")

# The marker line plus any decorators between it and the def, then the test function.
# Written as one regex so a marker can never be paired with the wrong test.
MARKER_RE = re.compile(
    r"@pytest\.mark\.be_gap([^\n]*)\n(?:@[^\n]*\n)*async def (test_[a-z0-9_]+)",
)

DOC_SUFFIX_EN = "_TEST_CASES.md"
DOC_SUFFIX_VI = "_TEST_CASES_vi.md"


# ── Module discovery ─────────────────────────────────────────────────────────
# A TC string starts with its module (utils/sync_registry.py derives it from the test
# function name), and each module documents itself in `{MODULE}_TEST_CASES.md` + the `_vi`
# translation. So the doc pair for a TC is DERIVED, never listed here:
#
#     PARTNER_UI_DASHBOARD_005  ->  PARTNER  ->  PARTNER_TEST_CASES.md  + _vi.md
#     PLANS_UI_VISIBLE_001      ->  PLANS    ->  PLANS_TEST_CASES.md    + _vi.md
#
# Adding a module means adding its docs pair, and every check below picks it up. Get the
# filename wrong and test_be_gap_module_has_a_docs_pair says which name it looked for.


def _known_modules() -> list[str]:
    """Module names that have an EN docs file, longest first (see _module_of)."""
    names = [p.name.removesuffix(DOC_SUFFIX_EN) for p in DOCS_DIR.glob(f"*{DOC_SUFFIX_EN}")]
    return sorted(names, key=len, reverse=True)


MODULES = _known_modules()


def _module_of(tc: str) -> str:
    """Module owning this TC string.

    Longest known module first, so a module whose name contains another module's name
    (say PARTNER and PARTNER_PORTAL) resolves to the right one. Unknown module → the
    first segment, which is what the "no docs pair" message needs to name.
    """
    for module in MODULES:
        if tc.startswith(f"{module}_"):
            return module
    return tc.split("_", 1)[0]


def _doc_pair(module: str) -> tuple[Path, Path]:
    return DOCS_DIR / f"{module}{DOC_SUFFIX_EN}", DOCS_DIR / f"{module}{DOC_SUFFIX_VI}"


# ── Scanners ─────────────────────────────────────────────────────────────────


def _scan_code() -> dict[str, list[str]]:
    """{TC_NAME: [bug ids named on its be_gap marker]} for every be_gap test."""
    found: dict[str, list[str]] = {}
    for path in sorted(TESTS_DIR.rglob("*.py")):
        for match in MARKER_RE.finditer(path.read_text(encoding="utf-8")):
            comment, func = match.group(1), match.group(2)
            found[func.upper().removeprefix("TEST_")] = BUG_ID_RE.findall(comment)
    return found


def _scan_tracker() -> tuple[dict[str, str], dict[str, list[str]]]:
    """({bug id: STATUS}, {TC_NAME: [bug ids]}) from the Bug_Tracker sheet.

    Rows whose Bug ID is not a real id are section banners, not bugs — skipped.
    Status cells can hold several lines (a bug closed in stages); the first line wins.
    """
    sheet = openpyxl.load_workbook(TRACKER, data_only=True)["Bug Tracker"]
    status: dict[str, str] = {}
    by_tc: dict[str, list[str]] = {}
    for row in range(2, sheet.max_row + 1):
        bug = str(sheet.cell(row, 1).value or "").strip()
        if not BUG_ID_RE.fullmatch(bug):
            continue
        raw = str(sheet.cell(row, 8).value or "").strip()
        status[bug] = (raw.splitlines() or [""])[0].strip().upper()
        tc = str(sheet.cell(row, 3).value or "").strip()
        if tc:
            by_tc.setdefault(tc, []).append(bug)
    return status, by_tc


def _tc_blocks(path: Path) -> dict[str, str]:
    """{TC_NAME: the whole `#### TC_NAME ...` block}; empty when the file is absent."""
    if not path.is_file():
        return {}
    blocks = re.split(r"^#### ", path.read_text(encoding="utf-8"), flags=re.M)[1:]
    return {block.split()[0].strip(): block for block in blocks}


CODE = _scan_code()
BUG_STATUS, TRACKER_BY_TC = _scan_tracker()

BE_GAP_TCS = sorted(CODE)
BE_GAP_MODULES = sorted({_module_of(tc) for tc in BE_GAP_TCS})
# (tc, doc path) for both languages of the TC's own module.
BE_GAP_IN_DOCS = [(tc, doc) for tc in BE_GAP_TCS for doc in _doc_pair(_module_of(tc))]
ALL_DOC_FILES = sorted(DOCS_DIR.glob(f"*{DOC_SUFFIX_EN}")) + sorted(
    DOCS_DIR.glob(f"*{DOC_SUFFIX_VI}")
)


def _ids(paths):
    """Readable parametrize ids: file name instead of the whole absolute path."""
    return [p.name for p in paths]


def test_the_scan_actually_found_the_markers():
    """A broken scan would make every check below vacuously green."""
    assert BE_GAP_TCS, "no be_gap markers found — the scan regex or the tests dir moved"
    assert BUG_STATUS, "no bugs read from the tracker — the sheet name or layout changed"
    assert MODULES, f"no '*{DOC_SUFFIX_EN}' docs found in {DOCS_DIR}"


# ── code marker → tracker ────────────────────────────────────────────────────


@pytest.mark.parametrize("tc", BE_GAP_TCS)
def test_marker_names_its_bug(tc):
    """Without an id on the marker there is nothing to follow when the bug closes."""
    assert CODE[tc], (
        f"{tc}: @pytest.mark.be_gap names no bug — add '# BUG-XXX-NNN: <why>' to the "
        "marker line, otherwise nobody can tell when this marker becomes stale"
    )


@pytest.mark.parametrize("tc", BE_GAP_TCS)
def test_marker_bug_exists_in_the_tracker(tc):
    unknown = [b for b in CODE[tc] if b not in BUG_STATUS]
    assert not unknown, (
        f"{tc}: bug id(s) {unknown} are not in Bug_Tracker.xlsx — a typo, or the bug was "
        "renamed there and the marker was not updated"
    )


@pytest.mark.parametrize("tc", BE_GAP_TCS)
def test_marker_bug_matches_the_tracker_row_for_this_tc(tc):
    """The tracker row that names this TC must name the same bug the code does."""
    expected = TRACKER_BY_TC.get(tc)
    assert expected, f"{tc}: is be_gap but no Bug_Tracker row names it"
    assert set(CODE[tc]) == set(expected), (
        f"{tc}: marker says {sorted(CODE[tc])} but the tracker row says {sorted(expected)}"
    )


@pytest.mark.parametrize("tc", BE_GAP_TCS)
def test_the_bug_is_still_open(tc):
    """A closed bug means the marker is stale: the TC belongs back inside the gate.

    This is the 3-week-latency check. Leaving it marked keeps a passing TC outside the
    merge gate, so a later regression in it would go unnoticed.
    """
    closed = [b for b in CODE[tc] if BUG_STATUS.get(b, "OPEN") != "OPEN"]
    assert not closed, (
        f"{tc}: {closed} is CLOSED in Bug_Tracker but the be_gap marker is still there. "
        "Re-run the TC; if it passes, drop the marker (code + both TEST_CASES docs) so "
        "the TC is back inside the merge gate"
    )


# ── code marker → docs ───────────────────────────────────────────────────────


@pytest.mark.parametrize("module", BE_GAP_MODULES)
def test_be_gap_module_has_a_docs_pair(module):
    """A module with be_gap TCs must document them, in both languages."""
    missing = [p.name for p in _doc_pair(module) if not p.is_file()]
    assert not missing, (
        f"module {module} has be_gap TC(s) but {missing} is missing from {DOCS_DIR}. "
        f"Every module documents itself in <MODULE>{DOC_SUFFIX_EN} plus the "
        f"<MODULE>{DOC_SUFFIX_VI} translation"
    )


@pytest.mark.parametrize(("tc", "doc"), BE_GAP_IN_DOCS, ids=lambda v: getattr(v, "name", v))
def test_doc_says_be_gap(tc, doc):
    """The docs must state that a be_gap TC sits outside the merge gate."""
    block = _tc_blocks(doc).get(tc)
    assert block is not None, f"{tc}: no '#### {tc}' entry in {doc.name}"
    assert "be_gap" in block, (
        f"{tc}: marked be_gap in code but {doc.name} never says so — a reader would "
        "believe this TC guards the merge gate when it is excluded from it"
    )


@pytest.mark.parametrize(("tc", "doc"), BE_GAP_IN_DOCS, ids=lambda v: getattr(v, "name", v))
def test_doc_names_the_same_bug(tc, doc):
    missing = set(CODE[tc]) - set(BUG_ID_RE.findall(_tc_blocks(doc).get(tc, "")))
    assert not missing, (
        f"{tc}: {doc.name} does not name {sorted(missing)} — the reader has to open the "
        "Excel tracker to find out which defect blocks this TC"
    )


# ── every bug id written in the docs must be a real one ──────────────────────
# Catches renames: BUG-028 → BUG-UI-008 was applied to one of the two docs only,
# because the two files had different surrounding wording.


@pytest.mark.parametrize("doc", ALL_DOC_FILES, ids=_ids(ALL_DOC_FILES))
def test_doc_bug_ids_all_exist_in_the_tracker(doc):
    unknown = sorted(set(BUG_ID_RE.findall(doc.read_text(encoding="utf-8"))) - set(BUG_STATUS))
    assert not unknown, f"{doc.name} references bug id(s) absent from Bug_Tracker.xlsx: {unknown}"


# ── EN ↔ VI parity, per module ───────────────────────────────────────────────
# The VI file is a translation, not an independent document: same TCs, same structure,
# same bug ids. Only the prose language differs. Drift here means one audience is
# reading a stale plan.


@pytest.mark.parametrize("module", MODULES)
def test_every_module_has_a_vi_translation(module):
    _, vi = _doc_pair(module)
    assert vi.is_file(), f"{module} has EN docs but no {vi.name}"


@pytest.mark.parametrize("module", MODULES)
def test_en_and_vi_cover_the_same_test_cases(module):
    en_path, vi_path = _doc_pair(module)
    en, vi = set(_tc_blocks(en_path)), set(_tc_blocks(vi_path))
    assert not en - vi, f"in {en_path.name} but not {vi_path.name}: {sorted(en - vi)}"
    assert not vi - en, f"in {vi_path.name} but not {en_path.name}: {sorted(vi - en)}"


@pytest.mark.parametrize("module", MODULES)
def test_en_and_vi_have_the_same_heading_structure(module):
    def headings(path: Path) -> int:
        if not path.is_file():
            return -1
        return sum(
            1
            for line in path.read_text(encoding="utf-8").splitlines()
            if re.match(r"#{1,4} ", line)
        )

    en_path, vi_path = _doc_pair(module)
    en, vi = headings(en_path), headings(vi_path)
    assert en == vi, (
        f"heading count differs: {en_path.name}={en} {vi_path.name}={vi} "
        "— a section was added to one file only"
    )


@pytest.mark.parametrize("module", MODULES)
def test_en_and_vi_reference_the_same_bugs(module):
    def bugs(path: Path) -> set[str]:
        return set(BUG_ID_RE.findall(path.read_text(encoding="utf-8"))) if path.is_file() else set()

    en_path, vi_path = _doc_pair(module)
    en, vi = bugs(en_path), bugs(vi_path)
    assert en == vi, (
        f"bug ids differ — only in {en_path.name}: {sorted(en - vi)}; only in "
        f"{vi_path.name}: {sorted(vi - en)}. A bug-id rename was applied to one file only"
    )
