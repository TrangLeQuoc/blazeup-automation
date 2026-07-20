#!/usr/bin/env python3
"""Regenerate runner/{domain}/registry.py from implemented test functions + Excel metadata.

Each domain gets its own registry file; runner/tc_registry.py then auto-merges
every runner/*/registry.py into one central TC_REGISTRY at import time.

How it works
------------
1. Scan tests/{domain}/**/*.py for functions named:
       test_partner_{ui|api}_{section}_{NNN}
   e.g. test_partner_ui_partner_portal_shell_001
        test_partner_api_auth_access_control_002
2. For each function, compute the numeric TC ID from the function name alone.
3. Look up title & priority from docs/{domain}/Partner_Platform_Test_Plan.xlsx (optional).
   Falls back to function docstring / "P2" default if Excel is unavailable.
4. Also pick up legacy test_tc* / test_tca* functions (legacy demo tests).
5. Write runner/{domain}/registry.py.

Only IMPLEMENTED test cases (i.e. functions that already exist in a test file)
are registered.  Unimplemented TCs produce no registry entry — just write the
function and run sync again.

File layout (both flat and section-subfolder are supported)
-----------------------------------------------------------
    tests/{domain}/ui/test_partner_ui_dashboard.py          <- flat, all Dashboard TCs
    tests/{domain}/ui/partner_portal_shell/
        test_partner_ui_partner_portal_shell.py             <- section file, all Shell TCs

Function naming convention
--------------------------
    test_partner_ui_partner_portal_shell_001   <- TC 1010101
    test_partner_ui_partner_portal_shell_002   <- TC 1010102
    test_partner_api_auth_access_control_001   <- TC 10101

TC ID Encoding
--------------
New-style:
    {type}{project}{module:02d}{section:02d}{seq:02d}

    Digit   Meaning         Values
    ------  --------------  -----------------------------------------------
    type    1 = UI, 0 = API
    project 1 = blazeup_partner, 2 = blazeup_admin   (see PROJECTS below)
    module  01-99   PER DOMAIN   (see MODULES below)
    section 01-99   PER MODULE   (see MODULE_SECTIONS below)
    seq     01-99   sequential within section

    The project digit makes IDs globally unique even when two domains reuse the
    SAME module name (e.g. both have a "PARTNERS" module) — so the merged
    registry never silently overwrites across projects.

    Examples
    PARTNERS_UI_DASHBOARD_001       (partner) -> 1 1 01 02 01 -> 11010201
    PARTNERS_API_..._001            (partner) -> 0 1 01 .. ..  -> 0101.... (7 digits)
    SHELL_UI_LOAD_TIME_PAGE_001     (admin)   -> 1 2 01 01 01 -> 12010101
    DASHBOARD_UI_VISIBLE_001        (admin)   -> 1 2 02 01 01 -> 12020101

    UI IDs are 8 digits (>= 11_000_000); API IDs are <= 7 digits.
    -> No collision between UI and API, nor across projects.

Legacy-style (legacy demo tests, preserved for backward-compat):
    test_tc01_*   -> 1001  (UI)
    test_tca01_*  ->    1  (API)

Usage
-----
    python utils/sync_registry.py                          # sync the blazeup registry
    python utils/sync_registry.py --table                  # print ID reference table
"""

import ast
import importlib.util
import re
import sys
from pathlib import Path

import yaml

# openpyxl is optional — only needed to read the Excel test plan. Detect whether
# it's installed without importing it at module scope (the real import happens
# inside _sync_domain, where it's actually used).
_HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None

PROJECT_ROOT = Path(__file__).resolve().parent.parent
TESTS_DIR = PROJECT_ROOT / "tests"  # tests/{domain}/api|ui/...
RUNNER_DIR = PROJECT_ROOT / "runner"  # runner/{domain}/registry.py output
CONFIG_DIR = PROJECT_ROOT / "config"  # config/{domain}/config.yaml input

# ---------------------------------------------------------------------------
# TC ID numbering — loaded from each project's config/<domain>/config.yaml,
# NOT hard-coded here, so this shared file stays project-agnostic. Add a new
# project by dropping a config/<domain>/config.yaml; no edits to this file.
#
# Each config.yaml provides:
#     project_number: <int>             # leading PROJECT digit in every TC ID
#     modules:
#       <MODULE>:
#         number: <int>                 # the module's 2-digit slot
#         ui:  {FEATURE: <int>, ...}    # section slots for UI tests
#         api: {FEATURE: <int>, ...}    # section slots for API tests
#     excel:                            # optional — supplies titles/priority
#       sheet: "<Sheet Name>"
#       module: <CODE_MODULE_NAME>      # module these rows map to
#       excel_prefix: <LEGACY_PREFIX>   # TestcaseId prefix as written in xlsx
#
# Assembled into the module-level dicts the rest of this file consumes:
#     PROJECTS[domain]                       -> project number
#     MODULES[domain][MODULE]                -> module number
#     MODULE_SECTIONS[domain][MODULE][TYPE]  -> {FEATURE: number}
#     EXCEL_SHEETS[sheet] -> {domain, module, excel_prefix}
#
# The project digit makes IDs globally unique even when two domains reuse the
# same module name (e.g. both have a "PARTNERS" module).
# ---------------------------------------------------------------------------


def _load_domain_configs() -> tuple[dict, dict, dict, dict]:
    """Read every config/<domain>/config.yaml into the numbering dicts."""
    projects: dict[str, int] = {}
    modules: dict[str, dict[str, int]] = {}
    module_sections: dict[str, dict[str, dict[str, dict[str, int]]]] = {}
    excel_sheets: dict[str, dict[str, str]] = {}

    for cfg_path in sorted(CONFIG_DIR.glob("*/config.yaml")):
        domain = cfg_path.parent.name
        try:
            data = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
        except (OSError, yaml.YAMLError) as exc:
            print(f"  [WARN] Could not read {cfg_path}: {exc}")
            continue

        # Project digit (accept legacy key 'module_number' as a fallback).
        project_num = data.get("project_number", data.get("module_number"))
        if project_num is None:
            continue
        projects[domain] = int(project_num)

        modules[domain] = {}
        module_sections[domain] = {}
        for mod_name, mod_cfg in (data.get("modules") or {}).items():
            mod_cfg = mod_cfg or {}
            if "number" not in mod_cfg:
                print(f"  [WARN] {domain}: module '{mod_name}' has no 'number' — skipped")
                continue
            modules[domain][mod_name] = int(mod_cfg["number"])
            sections: dict[str, dict[str, int]] = {}
            for tc_type in ("ui", "api"):
                feats = mod_cfg.get(tc_type)
                if feats:
                    sections[tc_type.upper()] = {k: int(v) for k, v in feats.items()}
            module_sections[domain][mod_name] = sections

        # Optional Excel sheet mapping for title/priority lookup.
        excel = data.get("excel") or {}
        sheet = excel.get("sheet")
        sheet_module = excel.get("module")
        if sheet and sheet_module:
            excel_sheets[sheet] = {
                "domain": domain,
                "module": sheet_module,
                "excel_prefix": excel.get("excel_prefix", sheet_module),
            }

    return projects, modules, module_sections, excel_sheets


PROJECTS, MODULES, MODULE_SECTIONS, EXCEL_SHEETS = _load_domain_configs()

# ---------------------------------------------------------------------------
# Registry file template
# ---------------------------------------------------------------------------
# Shared TestCase dataclass (written once to registry_modules/_base.py). Not a
# .format() template — braces are literal.
_BASE_TEMPLATE = '''\
"""Shared TestCase dataclass for the per-module registry files. (AUTO-GENERATED — do not edit)"""

from dataclasses import dataclass, field
from typing import Literal


@dataclass(frozen=True)
class TestCase:
    """Metadata for a single automation test case."""

    tc_id:     int
    tc_string: str                        # Excel TestcaseId  e.g. PARTNER_UI_DASHBOARD_001
    type:      Literal["api", "ui"]
    module:    str
    title:     str
    test_path: str
    test_func: str
    markers:   list[str] = field(default_factory=list)
    priority:  Literal["P1", "P2", "P3"] = "P2"

    @property
    def node_id(self) -> str:
        return f"{self.test_path}::{self.test_func}"
'''

# One file per top-level module (registry_modules/<module>.py). .format(module, domain, items).
_MODULE_TEMPLATE = '''\
"""TC registry — module `{module}`. (AUTO-GENERATED — do not edit.)

One file per top-level module so per-module PRs don't collide. Merged into the
domain registry by runner/{domain}/registry.py.
"""

from runner.{domain}.registry_modules._base import TestCase

TC_REGISTRY: dict[int, TestCase] = {{
{items}
}}
'''

# Domain aggregator (registry.py) — globs + merges every registry_modules/*.py.
# .format(domain). Stays a FILE so runner/tc_registry.py discovers it unchanged.
_AGG_TEMPLATE = '''\
"""Registry for the `{domain}` domain — merges runner/{domain}/registry_modules/*.py.
(AUTO-GENERATED — do not edit. Add TCs by writing tests + running utils/sync_registry.py.)

Each top-level module has its own file under registry_modules/ so per-module PRs
don't collide; this file globs + merges them into one TC_REGISTRY.
"""

import importlib
import pkgutil
from pathlib import Path

from runner.{domain}.registry_modules._base import TestCase

TC_REGISTRY: dict[int, TestCase] = {{}}

_pkg_path = Path(__file__).parent / "registry_modules"
for _m in pkgutil.iter_modules([str(_pkg_path)]):
    if _m.name.startswith("_"):
        continue
    _mod = importlib.import_module(f"runner.{domain}.registry_modules.{{_m.name}}")
    TC_REGISTRY.update(getattr(_mod, "TC_REGISTRY", {{}}))


def get_tc(tc_id: int) -> TestCase:
    if tc_id not in TC_REGISTRY:
        raise KeyError(f"TC {{tc_id}} does not exist in the registry")
    return TC_REGISTRY[tc_id]


def validate_registry() -> None:
    """Verify that all registered test functions exist (optional utility)."""
    for tc in TC_REGISTRY.values():
        if not Path(tc.test_path).exists():
            print(f"Warning: Test file {{tc.test_path}} missing for TC {{tc.tc_id}}")


def list_by_module(module: str) -> list[TestCase]:
    return [tc for tc in TC_REGISTRY.values() if tc.module == module]


def list_by_type(tc_type: str) -> list[TestCase]:
    return [tc for tc in TC_REGISTRY.values() if tc.type == tc_type]


def list_by_marker(marker: str) -> list[TestCase]:
    return [tc for tc in TC_REGISTRY.values() if marker in tc.markers]
'''


# ---------------------------------------------------------------------------
# ID computation
# ---------------------------------------------------------------------------


def tc_id_from_string(tc_string_id: str, domain: str) -> int:
    """Convert a TC string ID (within a domain) to its numeric equivalent.

    ID format: {type}{project}{module:02d}{section:02d}{seq:02d}
        type    1 = UI, 0 = API
        project PROJECTS[domain]  (partner=1, admin=2)

    PARTNERS_UI_DASHBOARD_001 (blazeup_partner) -> 1 1 01 02 01 -> 11010201
    PARTNERS_API_..._001       (blazeup_partner) -> 0 1 01 ..    -> 0101.... (7 digits)
    SHELL_UI_LOAD_TIME_PAGE_001 (blazeup_admin)  -> 1 2 01 01 01 -> 12010101

    Returns 0 on failure (unknown project / module / section, bad seq).
    """
    parts = tc_string_id.strip().split("_")
    # Minimum structure: MODULE_TYPE_SECTION_SEQ -> 4 parts
    if len(parts) < 4:
        return 0

    module_name = parts[0]  # PARTNERS | SHELL | DASHBOARD | ...
    tc_type = parts[1]  # UI | API
    seq_str = parts[-1]  # 001
    section_key = "_".join(parts[2:-1])  # DASHBOARD | AUTH_ACCESS_CONTROL | ...

    try:
        seq = int(seq_str)
    except ValueError:
        return 0
    if seq < 1 or seq > 99:
        return 0

    project_num = PROJECTS.get(domain, 0)
    module_num = MODULES.get(domain, {}).get(module_name, 0)
    sections = MODULE_SECTIONS.get(domain, {}).get(module_name, {}).get(tc_type, {})
    section_num = sections.get(section_key, 0)

    if project_num == 0 or module_num == 0 or section_num == 0:
        return 0

    type_digit = 1 if tc_type == "UI" else 0
    return int(f"{type_digit}{project_num}{module_num:02d}{section_num:02d}{seq:02d}")


def tc_string_from_id(tc_id: int) -> str:
    """Reverse-decode a numeric TC ID back to a human-readable label (debug only)."""
    s = f"{tc_id:08d}"
    type_char = s[0]
    project_num = int(s[1])
    module_num = int(s[2:4])
    section_num = int(s[4:6])
    seq = int(s[6:8])
    tc_type = "UI" if type_char == "1" else "API"
    domain = next((k for k, v in PROJECTS.items() if v == project_num), f"PROJ{project_num}")
    module_name = next(
        (k for k, v in MODULES.get(domain, {}).items() if v == module_num),
        f"MOD{module_num:02d}",
    )
    sections = MODULE_SECTIONS.get(domain, {}).get(module_name, {}).get(tc_type, {})
    section_name = next(
        (k for k, v in sections.items() if v == section_num), f"SEC{section_num:02d}"
    )
    return f"{module_name}_{tc_type}_{section_name}_{seq:02d}"


# ---------------------------------------------------------------------------
# Helper: function name -> TC string
# ---------------------------------------------------------------------------


def _func_name_to_tc_string(func_name: str, domain: str) -> str | None:
    """Convert a test function name to its TC string ID (within a domain).

    test_partners_ui_dashboard_001  -> PARTNERS_UI_DASHBOARD_001
    test_shell_ui_page_loads_001     -> SHELL_UI_PAGE_LOADS_001

    Returns None if the name doesn't match the expected pattern
    (must start with test_, end with _NNN where NNN is exactly 3 digits).
    """
    if not func_name.startswith("test_"):
        return None
    body = func_name[5:]  # strip "test_"
    parts = body.split("_")
    # Must end with a 3-digit sequence number
    if len(parts) < 4 or not re.match(r"^\d{3}$", parts[-1]):
        return None
    # First part must be a module key registered for THIS domain (e.g. "shell",
    # "dashboard", "partners"). Module names are namespaced per domain so the
    # same name can mean different things in blazeup_admin vs blazeup_partner.
    # This also prevents legacy functions like test_tca02_..._401 from matching.
    if parts[0].upper() not in MODULES.get(domain, {}):
        return None
    return body.upper()  # PARTNERS_UI_DASHBOARD_001


def _extract_markers(node: ast.FunctionDef | ast.AsyncFunctionDef) -> list[str]:
    """Extract @pytest.mark.X decorator names from an AST function node."""
    markers = []
    for dec in node.decorator_list:
        if (
            isinstance(dec, ast.Attribute)
            and isinstance(dec.value, ast.Attribute)
            and dec.value.attr == "mark"
        ):
            markers.append(dec.attr)
    return markers


# ---------------------------------------------------------------------------
# Partner Platform test scanner (new-style naming)
# ---------------------------------------------------------------------------


def _get_test_files(domain: str | None = None) -> list[Path]:
    """Return all test_*.py files.

    If domain is given (e.g. "blazeup_admin", "blazeup_partner"), scan only tests/{domain}/.
    Otherwise scan tests/*/ for all domains (auto-discover).
    """
    if domain:
        domain_dir = TESTS_DIR / domain
        return sorted(domain_dir.rglob("test_*.py")) if domain_dir.exists() else []
    # Auto-discover: any subfolder of tests/ that contains test files
    return sorted(
        py_file
        for domain_dir in sorted(TESTS_DIR.iterdir())
        if domain_dir.is_dir()
        for py_file in domain_dir.rglob("test_*.py")
    )


def scan_implemented_tcs(excel_lookup: dict[str, dict], domain: str | None = None) -> list[dict]:
    """Scan test files for implemented Partner Platform test functions.

    Discovers functions named:
        test_partner_{ui|api}_{section}_{NNN}

    A single file may contain many TC functions — all are registered.
    Title/priority come from Excel; docstring and P2 are fallbacks.
    """
    results: list[dict] = []
    seen_ids: set[int] = set()

    _all_test_files = _get_test_files(domain)

    for py_file in _all_test_files:
        rel_path = py_file.relative_to(PROJECT_ROOT).as_posix()

        try:
            source = py_file.read_text(encoding="utf-8")
            tree = ast.parse(source)
        except (OSError, SyntaxError):
            continue

        for node in ast.walk(tree):
            if not isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
                continue

            tc_string = _func_name_to_tc_string(node.name, domain)
            if tc_string is None:
                continue  # not a module TC function for this domain

            tc_id = tc_id_from_string(tc_string, domain)
            if tc_id == 0:
                # Function name matches pattern but section/module not registered
                print(
                    f"  [WARN] Unknown section in {py_file.name}: {node.name} "
                    f"(TC string: {tc_string})"
                )
                continue

            if tc_id in seen_ids:
                print(
                    f"  [ERROR] Duplicate TC ID {tc_id} ({tc_string}) "
                    f"in {rel_path} — already registered from another file"
                )
                continue
            seen_ids.add(tc_id)

            parts = tc_string.split("_")
            tc_type = parts[1].lower()  # ui / api
            module = parts[0].lower()  # partner

            # Metadata: Excel first, then docstring, then defaults
            meta = excel_lookup.get(tc_string, {})
            doc_line = (ast.get_docstring(node) or "").split("\n")[0]
            # Strip "TC_STRING: " prefix from docstring if present
            if ": " in doc_line:
                doc_line = doc_line.split(": ", 1)[-1]

            title = meta.get("title") or doc_line or "No Title"
            priority = meta.get("priority", "P2")
            markers = _extract_markers(node)

            results.append(
                {
                    "id": tc_id,
                    "type": tc_type,
                    "module": module,
                    "title": title.replace('"', "'"),
                    "priority": priority,
                    "path": rel_path,
                    "func": node.name,
                    "markers": markers,
                    "source": "scan",
                    "tc_string": tc_string,
                }
            )

    return results


# ---------------------------------------------------------------------------
# Legacy demo scanner (old test_tc* / test_tca* naming)
# ---------------------------------------------------------------------------


def scan_legacy_tcs(domain: str | None = None) -> list[dict]:
    """Scan test files for old-style test_tc* / test_tca* functions.

    These are legacy demo tests.

    IDs are assigned sequentially (1, 2, 3, …) sorted by:
        1. type  — api before ui
        2. file  — alphabetical within each type
        3. line  — function order within each file

    tc_string is set to "demo" to distinguish them from Partner Platform TCs.
    """
    raw: list[dict] = []

    _all_test_files = _get_test_files(domain)

    for py_file in _all_test_files:
        rel_path = py_file.relative_to(PROJECT_ROOT).as_posix()
        tc_type = "api" if "/api/" in rel_path else "ui"

        mod_match = re.search(r"test_(.*)_(?:api|ui)\.py$|test_(.*)\.py$", py_file.name)
        module_name = (mod_match.group(1) or mod_match.group(2)) if mod_match else "unknown"

        try:
            tree = ast.parse(py_file.read_text(encoding="utf-8"))
        except (OSError, SyntaxError):
            continue

        for node in ast.walk(tree):
            if not isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
                continue
            if not node.name.startswith("test_tc"):
                continue
            if not re.search(r"test_tc(a?)\d+", node.name):
                continue

            docstring = ast.get_docstring(node) or ""
            title = docstring.split("\n")[0].split(": ", 1)[-1] if docstring else "No Title"
            markers = _extract_markers(node)
            priority = "P1" if "smoke" in markers else "P2"

            raw.append(
                {
                    "type": tc_type,
                    "module": module_name,
                    "title": title.replace('"', "'"),
                    "priority": priority,
                    "path": rel_path,
                    "func": node.name,
                    "markers": markers,
                    "source": "legacy",
                    "tc_string": "demo",
                    "_lineno": node.lineno,
                }
            )

    # Sort: api first, then ui; within each group by file path then line number
    raw.sort(key=lambda x: (0 if x["type"] == "api" else 1, x["path"], x["_lineno"]))

    # Assign sequential IDs starting from 1
    results = []
    for seq, tc in enumerate(raw, start=1):
        entry = {k: v for k, v in tc.items() if k != "_lineno"}
        entry["id"] = seq
        results.append(entry)

    return results


# ---------------------------------------------------------------------------
# Registry diff helpers
# ---------------------------------------------------------------------------


def _parse_registry_snapshot(registry_file: Path) -> dict[int, str]:
    """Parse the existing registry into {tc_id: args_string} for diff comparison.

    Each line in the registry looks like:
        1010101: TestCase(1010101, "PARTNER_UI_...", "ui", ..., "P2"),
    We capture the args string inside TestCase(...) keyed by tc_id.
    Returns {} if the file doesn't exist yet (first-time sync).
    """
    if not registry_file.exists():
        return {}
    snapshot: dict[int, str] = {}
    for line in registry_file.read_text(encoding="utf-8").splitlines():
        m = re.match(r"\s+(\d+):\s*TestCase\((.+)\),\s*$", line)
        if m:
            snapshot[int(m.group(1))] = m.group(2)
    return snapshot


def _extract_fields(args_str: str) -> dict[str, str]:
    """Extract key fields from a raw TestCase args string.

    Double-quoted positional fields (0-indexed, markers use single quotes so they
    are not captured):
        0=tc_string  1=type  2=module  3=title  4=path  5=func  6=priority
    """
    quoted = re.findall(r'"([^"]*)"', args_str)
    return {
        "tc_string": quoted[0] if len(quoted) > 0 else "",
        "title": quoted[3] if len(quoted) > 3 else "",
        "path": quoted[4] if len(quoted) > 4 else "",
        "func": quoted[5] if len(quoted) > 5 else "",
        "priority": quoted[6] if len(quoted) > 6 else "",
    }


def _print_diff(
    old: dict[int, str],
    new: dict[int, str],
    new_tcs: list[dict],
) -> None:
    """Print a human-readable diff of registry changes."""
    added = sorted(set(new) - set(old))
    removed = sorted(set(old) - set(new))
    changed = sorted(tc_id for tc_id in set(new) & set(old) if new[tc_id] != old[tc_id])

    if not added and not removed and not changed:
        print("  No changes since last sync.")
        return

    tc_by_id = {tc["id"]: tc for tc in new_tcs}
    print(
        f"  Changes since last sync  "
        f"({len(added)} added, {len(removed)} removed, {len(changed)} changed):"
    )

    for tc_id in added:
        tc = tc_by_id.get(tc_id, {})
        title = tc.get("title", "")[:55]
        print(f'    +  TC {tc_id:<12} ADDED    {tc.get("func", "?")}  "{title}"')

    for tc_id in removed:
        old_f = _extract_fields(old[tc_id])
        print(f"    -  TC {tc_id:<12} REMOVED  {old_f.get('func', '?')}")

    for tc_id in changed:
        old_f = _extract_fields(old[tc_id])
        new_f = _extract_fields(new[tc_id])
        diffs: list[str] = []
        if old_f["title"] != new_f["title"]:
            diffs.append(f'title: "{old_f["title"][:30]}" -> "{new_f["title"][:30]}"')
        if old_f["priority"] != new_f["priority"]:
            diffs.append(f"priority: {old_f['priority']} -> {new_f['priority']}")
        if old_f["path"] != new_f["path"]:
            diffs.append("path changed")
        detail = "  |  ".join(diffs) if diffs else "content updated"
        tc = tc_by_id.get(tc_id, {})
        print(f"    ~  TC {tc_id:<12} CHANGED  {tc.get('func', '?')}  ({detail})")


# ---------------------------------------------------------------------------
# Main sync
# ---------------------------------------------------------------------------


def _sync_domain(domain: str) -> None:
    """Sync registry for a single domain → runner/{domain}/registry.py."""

    registry_file = RUNNER_DIR / domain / "registry.py"
    excel_file = PROJECT_ROOT / "docs" / domain / "Partner_Platform_Test_Plan.xlsx"

    print(f"\n{'=' * 55}")
    print(f"  Domain  : {domain}")
    print(f"  Tests   : {TESTS_DIR / domain}")
    print(f"  Excel   : {excel_file}")
    print(f"  Output  : {registry_file.relative_to(PROJECT_ROOT)}")
    print(f"{'=' * 55}")

    # Build Excel lookup for this domain's excel file
    excel_lookup: dict[str, dict] = {}
    if _HAS_OPENPYXL and excel_file.exists():
        import openpyxl  # type: ignore

        C_TC_STRING, C_TITLE, C_PRIORITY = 2, 3, 5
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        for sheet_name, cfg in EXCEL_SHEETS.items():
            # Only load sheets that belong to the domain being synced.
            if cfg["domain"] != domain:
                continue
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            excel_prefix = cfg["excel_prefix"]  # legacy prefix in the xlsx
            module_name = cfg["module"]  # code module name
            before = len(excel_lookup)
            for row in ws.iter_rows(min_row=13, values_only=True):
                tc_string = row[C_TC_STRING]
                if not tc_string:
                    continue
                tc_string = str(tc_string).strip()
                if not tc_string.startswith(f"{excel_prefix}_"):
                    continue
                # Re-key legacy Excel prefix -> code module name so the lookup
                # matches function-derived TC strings (xlsx stays untouched).
                if excel_prefix != module_name:
                    tc_string = module_name + tc_string[len(excel_prefix) :]
                title = str(row[C_TITLE] or "No Title").strip().replace('"', "'")
                priority = str(row[C_PRIORITY] or "P2").strip()
                excel_lookup[tc_string] = {"title": title, "priority": priority}
            print(f"[Excel] {len(excel_lookup) - before:4d} TCs from sheet '{sheet_name}'")
    else:
        print("[Excel] Skipped (file not found or openpyxl missing)")

    # Scan implemented TCs
    partner_tcs = scan_implemented_tcs(excel_lookup, domain=domain)
    partner_ids = {tc["id"] for tc in partner_tcs}
    print(f"[Scan]   {len(partner_tcs):4d} implemented module TCs found")

    # Scan legacy TCs (for SA domain)
    legacy_tcs = [tc for tc in scan_legacy_tcs(domain=domain) if tc["id"] not in partner_ids]
    print(f"[Legacy] {len(legacy_tcs):4d} legacy test_tc* TCs found")

    all_tcs = sorted(partner_tcs + legacy_tcs, key=lambda x: x["id"])

    # Check duplicates
    seen: dict[int, str] = {}
    for tc in all_tcs:
        if tc["id"] in seen:
            print(f"  [ERROR] Duplicate TC ID {tc['id']}: {seen[tc['id']]} vs {tc['func']}")
        else:
            seen[tc["id"]] = tc["func"]

    # Output layout: runner/{domain}/registry.py (aggregator, a FILE so tc_registry
    # discovers it unchanged) + runner/{domain}/registry_modules/<module>.py (one per
    # top-level module, so per-module PRs don't collide).
    domain_dir = RUNNER_DIR / domain
    domain_dir.mkdir(parents=True, exist_ok=True)
    (domain_dir / "__init__.py").touch(exist_ok=True)
    modules_dir = domain_dir / "registry_modules"
    modules_dir.mkdir(parents=True, exist_ok=True)
    (modules_dir / "__init__.py").touch(exist_ok=True)

    # Snapshot before overwrite (parse every existing per-module file for the diff)
    old_snapshot: dict[int, str] = {}
    for f in modules_dir.glob("*.py"):
        if not f.name.startswith("_"):
            old_snapshot.update(_parse_registry_snapshot(f))

    # Build per-module buckets (all_tcs is already sorted by id → stable order)
    buckets: dict[str, list[str]] = {}
    new_snapshot: dict[int, str] = {}
    for tc in all_tcs:
        tc_string = tc.get("tc_string", "")
        args_str = (
            f"{tc['id']}, "
            f'"{tc_string}", '
            f'"{tc["type"]}", '
            f'"{tc["module"]}", '
            f'"{tc["title"]}", '
            f'"{tc["path"]}", '
            f'"{tc["func"]}", '
            f"{tc['markers']}, "
            f'"{tc["priority"]}"'
        )
        new_snapshot[tc["id"]] = args_str
        buckets.setdefault(tc["module"], []).append(f"    {tc['id']}: TestCase({args_str}),")

    _wr = {"encoding": "utf-8", "newline": "\n"}  # LF (matches .gitattributes)
    (modules_dir / "_base.py").write_text(_BASE_TEMPLATE, **_wr)

    wanted = {"_base.py", "__init__.py"}
    for module, lines in sorted(buckets.items()):
        wanted.add(f"{module}.py")
        (modules_dir / f"{module}.py").write_text(
            _MODULE_TEMPLATE.format(module=module, domain=domain, items="\n".join(lines)), **_wr
        )
    # Remove stale per-module files (a module that no longer has any TC)
    for f in modules_dir.glob("*.py"):
        if f.name not in wanted:
            f.unlink()

    registry_file.write_text(_AGG_TEMPLATE.format(domain=domain), **_wr)

    print(
        f"\n  Total : {len(all_tcs):4d} TCs across {len(buckets)} module file(s) "
        f"in {modules_dir.relative_to(PROJECT_ROOT)}"
    )
    _print_diff(old_snapshot, new_snapshot, all_tcs)


def sync(domain: str | None = None) -> None:
    """Regenerate runner/{domain}/registry.py for one or all domains.

    domain=None  → sync all domains found in tests/*/
    domain="blazeup_admin"  → sync SA only  → runner/blazeup/registry.py
    """
    print("Syncing TC registry ...")
    print(f"  Tests root : {TESTS_DIR}")

    if domain:
        _sync_domain(domain)
    else:
        # Auto-discover all domains in tests/
        domains = sorted(
            d.name for d in TESTS_DIR.iterdir() if d.is_dir() and not d.name.startswith("_")
        )
        if not domains:
            print("  No domains found in tests/. Create tests/{domain}/ first.")
            return
        print(f"  Domains    : {', '.join(domains)}")
        for d in domains:
            _sync_domain(d)


# ---------------------------------------------------------------------------
# Reference table printer
# ---------------------------------------------------------------------------


def print_id_table() -> None:
    """Print the full domain -> module -> section -> numeric-prefix table."""
    print("TC ID Reference Table")
    print("=" * 65)
    for domain, modules in MODULES.items():
        project_num = PROJECTS.get(domain, 0)
        print(f"\n=== {domain} (project {project_num}) ===")
        for mod_name, mod_num in modules.items():
            print(f"\n  Module: {mod_name} ({mod_num:02d})")
            for tc_type, sections in MODULE_SECTIONS.get(domain, {}).get(mod_name, {}).items():
                type_digit = 1 if tc_type == "UI" else 0
                print(f"    {tc_type}:")
                for sec_name, sec_num in sections.items():
                    first_id = int(f"{type_digit}{project_num}{mod_num:02d}{sec_num:02d}01")
                    print(f"      {sec_num:02d}  {sec_name:<42s}  first ID: {first_id}")


if __name__ == "__main__":
    if "--table" in sys.argv:
        print_id_table()
    else:
        # Support: python utils/sync_registry.py --domain sa
        _domain_arg: str | None = None
        if "--domain" in sys.argv:
            _idx = sys.argv.index("--domain")
            if _idx + 1 < len(sys.argv):
                _domain_arg = sys.argv[_idx + 1]
        sync(domain=_domain_arg)
