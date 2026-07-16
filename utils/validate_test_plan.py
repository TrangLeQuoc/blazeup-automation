"""Test-plan linter — validate the Excel test plan against the column contract.

It READS the workbook and reports violations; it never modifies the file.
Exit code 0 = clean (warnings allowed), 1 = at least one ERROR, 2 = cannot read.

Usage::

    python utils/validate_test_plan.py                      # default domain (blazeup)
    python utils/validate_test_plan.py --file path/to.xlsx --sheet "Partner Platform"
    python utils/validate_test_plan.py --strict             # warnings also fail (exit 1)

Column contract (the "schema" this enforces)
---------------------------------------------
Required columns must exist; required cells must be non-empty; enum columns may
only hold allowed values; the "Test Case Name" string id (col C, e.g.
``PARTNER_API_DEAL_..._001``) must be unique + well-formed; negative-scoped TCs
should be typed Negative/Security; cols A/B must not duplicate; any TC that
already has automation (present in the registry) must be flagged Auto = YES.

Note on the two id columns: col C "Test Case Name" holds the *string* id that
matches the code registry (this is what we validate/cross-check); col D "Test
Case ID" holds the *numeric* automation id and is not linted here.
"""

import argparse
import contextlib
import re
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SHEET = "Partner Platform"

# ── Column contract (header names exactly as they appear on the sheet) ────────
# The string id lives in "Test Case Name" (col C); "Action" (col E) is the
# human-readable scenario. "Test Case ID" (col D) is the numeric automation id.
COL_ID = "Test Case Name"  # string id we validate + cross-check with the registry
COL_ACTION = "Action"  # scenario text, scanned for negative wording

REQUIRED_COLUMNS = [
    "Main Section",
    "Main Feature/Function",
    COL_ID,
    COL_ACTION,
    "Test Type",
    "Priority",
    "Severity",
    "Status",
    "Auto",
    "Automation Status",
]
REQUIRED_NONEMPTY = [COL_ID, COL_ACTION, "Test Type", "Priority", "Severity", "Status", "Owner"]

ENUMS: dict[str, set[str]] = {
    "Test Type": {"Functional", "Negative", "Security", "Performance", "Compliance"},
    "Priority": {"P1", "P2", "P3"},
    "Severity": {"Critical", "Major", "Minor"},
    "Status": {"NOT_STARTED", "IN_PROGRESS", "PASSED", "FAILED", "BLOCKED"},
    "Manual Status": {"NOT_STARTED", "IN_PROGRESS", "PASSED", "FAILED", "BLOCKED"},
    "Auto": {"YES", "NO"},
    "Automation Status": {"NOT_STARTED", "IN_PROGRESS", "PASSED", "FAILED", "BLOCKED"},
}

ID_RE = re.compile(r"^[A-Z][A-Z0-9]*_(UI|API)_[A-Z0-9_]+_\d{3}$")
NEG_RE = re.compile(
    r"\b(validate|invalid|reject|missing|malformed|unauthor|forbidden|denied|"
    r"block|400|401|403|409)\b",
    re.I,
)
BOILERPLATE_HINTS = ("Prepare API fixtures", "Validate the backend/API contract for")


class Report:
    """Collects ERROR / WARN findings keyed by row."""

    def __init__(self) -> None:
        self.errors: list[str] = []
        self.warnings: list[str] = []

    def error(self, row: int, tc: str, msg: str) -> None:
        self.errors.append(f"  ❌ row {row:>3}  {tc:<46} {msg}")

    def warn(self, row: int, tc: str, msg: str) -> None:
        self.warnings.append(f"  ⚠  row {row:>3}  {tc:<46} {msg}")


def _resolve_path(args: argparse.Namespace) -> Path:
    if args.file:
        return Path(args.file)
    return PROJECT_ROOT / "docs" / args.domain / "Partner_Platform_Test_Plan.xlsx"


def _find_header_row(ws) -> int:
    """Return the row index that holds the column headers.

    Anchored on the "Test Case ID" cell (a banner row above it may say
    "UI Test Cases", so we scan for the real header, not just row 1).
    """
    for r in range(1, 40):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(r, c).value or "").strip() == "Test Case ID":
                return r
    raise ValueError("Header row with 'Test Case ID' not found in the first 40 rows")


def _registry_tc_strings() -> set[str]:
    """Best-effort: TC strings that already have automation (from the registry)."""
    try:
        from runner.tc_registry import TC_REGISTRY
    except Exception:  # noqa: BLE001 — registry optional; skip the plan↔code check
        return set()
    out: set[str] = set()
    for tc in TC_REGISTRY.values():
        s = getattr(tc, "tc_string", None)
        if s and s != "demo":
            out.add(str(s))
    return out


def validate(path: Path, sheet: str) -> "Report | None":
    if not path.exists():
        print(f"ERROR: test plan not found: {path}", file=sys.stderr)
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        print(f"ERROR: sheet '{sheet}' not in {path.name} ({wb.sheetnames})", file=sys.stderr)
        return None
    ws = wb[sheet]
    hdr = _find_header_row(ws)
    cols = {
        str(ws.cell(hdr, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(hdr, c).value
    }

    rep = Report()

    # 0) required columns present
    missing_cols = [c for c in REQUIRED_COLUMNS if c not in cols]
    if missing_cols:
        rep.errors.append(f"  ❌ HEADER  missing required column(s): {missing_cols}")
        return rep  # without these, per-row checks are meaningless

    def val(r: int, name: str):
        c = cols.get(name)
        return ws.cell(r, c).value if c else None

    automated = _registry_tc_strings()
    seen_ids: dict[str, int] = {}

    for r in range(hdr + 1, ws.max_row + 1):
        tc = val(r, COL_ID)
        if not tc:
            continue  # section banner / blank row — not a test case
        tc = str(tc).strip()

        # 1) duplicate id
        if tc in seen_ids:
            rep.error(r, tc, f"duplicate Test Case Name (also at row {seen_ids[tc]})")
        else:
            seen_ids[tc] = r

        # 2) id format
        if not ID_RE.match(tc):
            rep.error(
                r, tc, "Test Case Name format invalid (expect <PREFIX>_(UI|API)_<MODULE>_NNN)"
            )

        # 3) required non-empty cells
        for f in REQUIRED_NONEMPTY:
            if f in cols and not str(val(r, f) or "").strip():
                rep.error(r, tc, f"required cell '{f}' is empty")

        # 4) enum validation
        for f, allowed in ENUMS.items():
            if f not in cols:
                continue
            v = val(r, f)
            if v is None or str(v).strip() == "":
                continue  # emptiness handled by required check / left as-is
            if str(v).strip() not in allowed:
                rep.error(r, tc, f"'{f}' = {v!r} not allowed (use one of {sorted(allowed)})")

        # 5) negative-scoped TC (by its Action wording) should be typed Negative/Security
        action = str(val(r, COL_ACTION) or "")
        ttype = str(val(r, "Test Type") or "").strip()
        if NEG_RE.search(action) and ttype not in ("Negative", "Security"):
            rep.warn(
                r,
                tc,
                f"Action looks negative but Test Type = {ttype!r} (consider Negative/Security)",
            )

        # 6) cols A/B must not duplicate
        sec = str(val(r, "Main Section") or "").strip()
        feat = str(val(r, "Main Feature/Function") or "").strip()
        if sec and feat and sec.lower() == feat.lower():
            rep.warn(r, tc, "Main Section == Main Feature/Function (cols A/B duplicated)")

        # 7) plan ↔ code: automated TC must be flagged Auto = YES
        if tc in automated and str(val(r, "Auto") or "").strip().upper() != "YES":
            rep.error(r, tc, "has automation (in registry) but Auto != YES")

        # 8) boilerplate description/step on an automated TC
        if tc in automated:
            blob = f"{val(r, 'Test Description') or ''} {val(r, 'Test Step') or ''}"
            if any(h in blob for h in BOILERPLATE_HINTS):
                rep.warn(r, tc, "Description/Test Step still looks like generic boilerplate")

    return rep


def main() -> int:
    # Windows consoles default to cp1252 and choke on the ✅/⚠/❌ glyphs — force UTF-8.
    for stream in (sys.stdout, sys.stderr):
        with contextlib.suppress(Exception):  # older/odd streams: fall back to default
            stream.reconfigure(encoding="utf-8")

    ap = argparse.ArgumentParser(description="Validate the Excel test plan against the schema.")
    ap.add_argument("--domain", default="blazeup", help="domain folder under docs/")
    ap.add_argument("--file", help="explicit path to the .xlsx (overrides --domain)")
    ap.add_argument("--sheet", default=DEFAULT_SHEET, help="worksheet name")
    ap.add_argument("--strict", action="store_true", help="treat warnings as failures too")
    args = ap.parse_args()

    path = _resolve_path(args)
    if not path.exists():
        # An explicit --file that's missing is a user mistake (fail). A domain that
        # simply has no plan is "nothing to lint" — not an error.
        if args.file:
            print(f"ERROR: --file not found: {path}", file=sys.stderr)
            return 2
        print(
            f"\nNo test plan for domain '{args.domain}' (expected {path.name}) — nothing to validate."
        )
        return 0

    rep = validate(path, args.sheet)
    if rep is None:
        return 2

    print(f"\nValidating {path.name}  (sheet '{args.sheet}')")
    if rep.errors:
        print(f"\n{len(rep.errors)} ERROR(s):")
        print("\n".join(rep.errors))
    if rep.warnings:
        print(f"\n{len(rep.warnings)} WARNING(s):")
        print("\n".join(rep.warnings))
    if not rep.errors and not rep.warnings:
        print("\n✅ clean — no issues found.")

    print(f"\n→ {len(rep.errors)} error(s), {len(rep.warnings)} warning(s)")
    if rep.errors or (args.strict and rep.warnings):
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
