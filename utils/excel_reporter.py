"""Export automation test results back into a copy of the test-plan Excel workbook.

Usage (called from runner/test_runner.py when --excel-report is active):

    from utils.excel_reporter import write_excel_report

    out_path = write_excel_report(tc_summaries, result_dir, excel_path)
    # out_path is None when nothing to write (e.g. only legacy "demo" TCs were run)

Excel column mapping (Partner_Platform_Test_Plan.xlsx, "Partner Platform" sheet)
----------------------------------------------------------------------------------
Col C  (3)  Test Case Name     — lookup key  (e.g. PARTNER_API_AUTH_ACCESS_CONTROL_001)
Col I  (9)  Status             — formula, auto-recomputes in Excel — DO NOT WRITE
Col J  (10) Manual Status      — not touched by automation
Col K  (11) Auto               — set to "YES" when automation covers the TC
Col L  (12) Automation Status  — written with PASSED / FAILED / BLOCKED / NOT_STARTED

Status values accepted by the workbook formulas
------------------------------------------------
  PASSED  /  FAILED  /  IN_PROGRESS  /  BLOCKED  /  NOT_STARTED
"""

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import yaml
from loguru import logger

_PROJECT_ROOT = Path(__file__).resolve().parents[1]

# ---------------------------------------------------------------------------
# Excel mapping is read PER-DOMAIN from config/<domain>/config.yaml (block
# `excel:`), so covering a new sheet (e.g. "Tenant") or shifting a column is a
# YAML edit — no code change. These are the fallbacks when the block/key is absent.
# ---------------------------------------------------------------------------

_DEFAULT_SHEETS = ["Partner Platform"]
_DEFAULT_COL_TC_STRING = 3  # C: Test Case Name (lookup key)
_DEFAULT_COL_AUTO_FLAG = 11  # K: Auto ("YES" / "NO")
_DEFAULT_COL_AUTO_STATUS = 12  # L: Automation Status
_DEFAULT_DATA_START_ROW = 13  # first row that contains test-case data


def _load_excel_cfg(excel_path: Path) -> dict:
    """Read the reporter's Excel mapping from config/<domain>/config.yaml.

    Domain is inferred from the test-plan path (``docs/<domain>/<file>.xlsx``).
    Every key falls back to the module defaults above, so a missing/partial
    ``excel:`` block still works. To cover a new sheet, add its name to
    ``excel.sheets`` in the domain's config.yaml — nothing here changes.
    """
    domain = excel_path.parent.name
    cfg_path = _PROJECT_ROOT / "config" / domain / "config.yaml"
    excel: dict = {}
    try:
        data = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
        excel = data.get("excel") or {}
    except Exception as exc:  # noqa: BLE001 — missing/bad config → use defaults
        logger.debug("Excel reporter: could not read {} ({}) — using defaults", cfg_path, exc)
    return {
        "sheets": excel.get("sheets") or _DEFAULT_SHEETS,
        "col_tc_string": excel.get("col_tc_string", _DEFAULT_COL_TC_STRING),
        "col_auto_flag": excel.get("col_auto_flag", _DEFAULT_COL_AUTO_FLAG),
        "col_auto_status": excel.get("col_auto_status", _DEFAULT_COL_AUTO_STATUS),
        "data_start_row": excel.get("data_start_row", _DEFAULT_DATA_START_ROW),
    }


# ---------------------------------------------------------------------------
# pytest outcome → Excel automation status
# ---------------------------------------------------------------------------

_OUTCOME_MAP: dict[str, str] = {
    "PASSED": "PASSED",
    "FAILED": "FAILED",
    "ERROR": "FAILED",
    "BLOCKED": "BLOCKED",
    "SKIPPED": "NOT_STARTED",
    "MISSING": "NOT_STARTED",
    "UNKNOWN": "NOT_STARTED",
}


def _outcome_from_summary(s: dict) -> str:
    """Convert a tc_summary dict into an Excel automation status string."""
    if s["failed"] > 0:
        return "FAILED"
    if s.get("blocked", 0) > 0:
        return "BLOCKED"
    if s["passed"] > 0:
        return "PASSED"
    # skipped / missing
    return "NOT_STARTED"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def write_excel_report(
    tc_summaries: list[dict],
    result_dir: Path,
    excel_path: Path,
) -> "Path | None":
    """Copy *excel_path*, fill automation results, save to *result_dir*.

    Args:
        tc_summaries:  List of summary dicts produced by
                       ``_build_tc_summaries_from_rows()`` / ``_build_tc_summaries_from_perf()``.
                       Each dict has keys: tc_id, passed, failed, skipped, …
        result_dir:    Run result directory (e.g. results/run_20260526_113058).
                       The output file is written here.
        excel_path:    Source test-plan Excel file
                       (Partner_Platform_Test_Plan.xlsx at project root).

    Returns:
        Path to the generated ``.xlsx`` file, or ``None`` when nothing was
        written (no matchable TCs in the run, or source file missing).
    """
    if not excel_path.exists():
        logger.warning("Excel template not found: {} — skipping Excel report", excel_path)
        return None

    # ── Build tc_string → outcome lookup from summaries ───────────────────
    try:
        from runner.tc_registry import TC_REGISTRY
    except ModuleNotFoundError:
        from tc_registry import TC_REGISTRY  # type: ignore[no-redef]

    result_lookup: dict[str, str] = {}
    for s in tc_summaries:
        tc = TC_REGISTRY.get(s["tc_id"])
        if tc is None:
            continue
        if not tc.tc_string or tc.tc_string == "demo":
            # Legacy / demo TCs have no Excel row — skip
            continue
        result_lookup[tc.tc_string] = _outcome_from_summary(s)

    if not result_lookup:
        logger.debug("Excel report: no new-style TCs in this run — skipping")
        return None

    # ── Copy the workbook ──────────────────────────────────────────────────
    # Reuse the run dir's timestamp (run_YYYYMMDD_HHMMSS) so the Excel filename
    # matches the folder. Fall back to now() if the dir name isn't a run_* folder.
    run_name = result_dir.name
    timestamp = (
        run_name[len("run_") :]
        if run_name.startswith("run_")
        else datetime.now().strftime("%Y%m%d_%H%M%S")
    )
    out_name = f"{excel_path.stem}_result_{timestamp}.xlsx"
    out_path = result_dir / out_name
    shutil.copy2(excel_path, out_path)

    # ── Open and update (sheets + columns from config/<domain>/config.yaml) ──
    cfg = _load_excel_cfg(excel_path)
    wb = openpyxl.load_workbook(out_path)
    updated = 0

    for sheet_name in cfg["sheets"]:
        if sheet_name not in wb.sheetnames:
            logger.debug("Sheet '{}' not in workbook — skipping", sheet_name)
            continue

        ws = wb[sheet_name]
        for row_idx in range(cfg["data_start_row"], ws.max_row + 1):
            tc_str = ws.cell(row=row_idx, column=cfg["col_tc_string"]).value
            if not tc_str:
                continue
            tc_str = str(tc_str).strip()
            if tc_str not in result_lookup:
                continue

            outcome = result_lookup[tc_str]
            ws.cell(row=row_idx, column=cfg["col_auto_flag"]).value = "YES"
            ws.cell(row=row_idx, column=cfg["col_auto_status"]).value = outcome
            logger.debug("  {} → Automation Status = {}", tc_str, outcome)
            updated += 1

    if updated == 0:
        # Nothing matched a workbook row — don't leave an unchanged copy behind.
        # Discard the file we copied and report nothing, even though Excel
        # reporting was requested (REPORT_EXCEL=True).
        wb.close()
        out_path.unlink(missing_ok=True)
        logger.warning(
            "Excel report: none of the {} tc_string(s) matched any row in the workbook "
            "- no result file written",
            len(result_lookup),
        )
        return None

    wb.save(out_path)
    logger.info("Excel report saved: {} ({} TC(s) updated)", out_path.name, updated)
    return out_path
