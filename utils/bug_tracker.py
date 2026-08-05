"""Reconcile triaged failures against the Bug Tracker (docs/blazeup/Bug_Tracker.xlsx).

Deterministic — keyed by **Test Case ID**, never the LLM. For each real defect
(triage category ``app_bug`` — API or UI), decide:

* NEW        — the TC has no row yet → append one (local runs only).
* KNOWN OPEN — a row exists and is still open → just report the Bug ID.
* REOPEN     — a row exists but is Closed and the test failed again → report ⚠.

The mirror of REOPEN is reported too, keyed by the run's PASSED test-case ids:

* RESOLVED   — a row exists and is still open, but the TC PASSED this run → the
               defect is no longer reproducing → report ✅ and suggest closing it.
               Suggest-only: the tracker's Status is NEVER auto-changed (a single
               pass can be flaky/env — a human decides the close).

Writes happen only on **local** runs (``CI`` env unset) and only to this separate
file — never the test plan. On CI it is propose-only (report, no file change).
"""

import datetime
import os
import re
from dataclasses import dataclass, field
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TRACKER = _PROJECT_ROOT / "docs" / "blazeup" / "Bug_Tracker.xlsx"
_SHEET = "Bug Tracker"
# Status stems that mean "no longer open". Matched by PREFIX (case-insensitive) so
# human variants all count: CLOSE / Closed / closing, Fix / Fixed, Resolve / Resolved,
# Verify / Verified, Done, Won't Fix / WontFix, Complete / Completed.
_CLOSED_STEMS = ("close", "fix", "resolv", "verif", "done", "wontfix", "won't", "complete")


def _is_closed(status: str | None) -> bool:
    """True when a tracker Status means the bug is no longer open (prefix match)."""
    s = (status or "").strip().lower()
    return any(s.startswith(stem) for stem in _CLOSED_STEMS)


@dataclass
class BugRow:
    tc_id: int
    bug_id: str
    status: str
    evidence: str = ""  # the assertion/evidence recorded when the bug was opened


@dataclass
class Reconciliation:
    new: list[dict] = field(default_factory=list)
    known_open: list[BugRow] = field(default_factory=list)
    reopen: list[BugRow] = field(default_factory=list)
    resolved: list[BugRow] = field(default_factory=list)
    written: int = 0
    wrote_to_file: bool = False
    tracker: Path = DEFAULT_TRACKER


def _tc_int(tc: str) -> int | None:
    m = re.search(r"\d+", tc or "")
    return int(m.group()) if m else None


# Normalize a failure message to a stable signature (mongo ids / numbers / quoted
# values masked) so the SAME root cause collapses to one string. Mirrors
# ai_triage._signature — used to tell whether a re-failure is the SAME defect as a
# closed bug (→ reopen) or a DIFFERENT one (→ open a new bug).
_ID_RE = re.compile(r"\b[0-9a-fA-F]{24}\b")
_QUOTED_RE = re.compile(r"(['\"]).*?\1")
_NUM_RE = re.compile(r"\d+")


def _sig(text: str) -> str:
    core = (text or "").split(" -- ", 1)[-1]
    if "AssertionError:" in core:
        core = core.split("AssertionError:", 1)[-1]
    s = _ID_RE.sub("<id>", core.strip())
    s = _QUOTED_RE.sub("'<v>'", s)
    s = _NUM_RE.sub("<n>", s)
    return s.lower()[:120]


# Most assertions read "Expected <X>, got <Y>" / "Expected <X> got <Y>" — split that
# into the tracker's Expected/Actual columns. When there is no such pattern (e.g. a UI
# content-load banner), Actual = the observed evidence and Expected is left blank for a
# human to fill (never fabricated).
_EXP_GOT_RE = re.compile(r"[Ee]xpected\s+(.+?)[,;]?\s+got\s+(.+)", re.DOTALL)


def _expected_actual(evidence: str) -> tuple[str, str]:
    ev = (evidence or "").strip()
    m = _EXP_GOT_RE.search(ev)
    if m:
        return m.group(1).strip()[:250], m.group(2).strip()[:250]
    return "", ev[:250]


def _registry_lookup() -> dict[int, object]:
    """Best-effort {tc_id: TestCase} from the code registry (empty on any failure)."""
    try:
        from runner.tc_registry import TC_REGISTRY

        return dict(TC_REGISTRY)
    except Exception:  # noqa: BLE001 — enrichment is optional; a stub row still works
        return {}


def _feature_from_tc_string(tc_string: str) -> str:
    """Derive the functional feature from a TC string.

    ``PARTNER_API_AUTH_ACCESS_CONTROL_003`` -> ``AUTH_ACCESS_CONTROL``.
    """
    m = re.match(r"^[A-Z0-9]+_(?:API|UI)_(.+)_\d+$", tc_string or "")
    return m.group(1) if m else ""


def load_tracker(path: Path) -> dict[int, BugRow]:
    """Return ``{tc_id: BugRow}`` keyed by Test Case ID (empty if no file/sheet)."""
    if not path.exists():
        return {}
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if _SHEET not in wb.sheetnames:
        return {}
    ws = wb[_SHEET]
    hdr = {str(c.value).strip().lower(): i for i, c in enumerate(ws[1], 1) if c.value}
    ci_bug, ci_tc, ci_st = hdr.get("bug id"), hdr.get("test case id"), hdr.get("status")
    ci_ev = hdr.get("evidence / assertion") or hdr.get("summary")
    out: dict[int, BugRow] = {}
    if not (ci_bug and ci_tc and ci_st):
        return out
    for row in ws.iter_rows(min_row=2):
        tc = _tc_int(str(row[ci_tc - 1].value or ""))
        if tc is None:
            continue
        out[tc] = BugRow(
            tc_id=tc,
            bug_id=str(row[ci_bug - 1].value or "").strip(),
            status=str(row[ci_st - 1].value or "").strip(),
            evidence=(str(row[ci_ev - 1].value or "").strip() if ci_ev else ""),
        )
    return out


# ── Section-aware Bug IDs ─────────────────────────────────────────────────────
# The tracker is split into an **API** section and a **UI** section, each opened by
# a bold header row (col A == "API" / "UI", the other columns blank). Bug IDs are
# prefixed + numbered per section: BUG-API-001.. and BUG-UI-001.. . New bugs are
# inserted at the END of their own section so ids stay stable and the two lists
# never interleave.
_SECTIONS = ("API", "UI")


def _is_ui(tc_id: int, tc_string: str = "") -> bool:
    """UI vs API classification for a test case (drives the Bug ID section)."""
    if "_UI_" in (tc_string or ""):
        return True
    if "_API_" in (tc_string or ""):
        return False
    # Fallback by id magnitude: UI ids are 8-digit and start with 1 (e.g. 12060515);
    # API ids are 7-digit and start with 2 (e.g. 2060234).
    return len(str(tc_id)) >= 8 and str(tc_id).startswith("1")


def _section_for(tc_id: int, tc_string: str = "") -> str:
    return "UI" if _is_ui(tc_id, tc_string) else "API"


def _next_seq(existing: dict[int, BugRow], section: str) -> int:
    """Highest existing sequence for BUG-<section>-NNN (0 if none)."""
    pat = re.compile(rf"BUG-{section}-0*(\d+)", re.IGNORECASE)
    mx = 0
    for r in existing.values():
        m = pat.match(r.bug_id or "")
        if m:
            mx = max(mx, int(m.group(1)))
    return mx


def _find_section_rows(ws) -> dict[str, int]:
    """{'API': header_row, 'UI': header_row} by scanning col A for the labels."""
    out: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or "").strip().upper()
        if label in _SECTIONS and (ws.cell(r, 2).value in (None, "")):
            out[label] = r
    return out


def _last_bug_row(ws, start: int, end: int) -> int:
    """Last row in (start, end) holding a Bug ID; else `start` (the section header)."""
    last = start
    for r in range(start + 1, end):
        if str(ws.cell(r, 1).value or "").upper().startswith("BUG-"):
            last = r
    return last


def _write_row(ws, idx: int, row: dict, hdr: dict[str, int]) -> None:
    for key, val in row.items():
        if key.startswith("_") or key not in hdr:  # skip transient hints (_section, …)
            continue
        cell = ws.cell(idx, hdr[key], val)
        # Test Case ID is a plain integer id, not a quantity — force the "0" number
        # format so it never renders with thousands separators (e.g. 12,060,102).
        if key == "Test Case ID":
            cell.number_format = "0"


def _append_rows(path: Path, rows: list[dict]) -> int:
    """Insert each new bug at the end of its section (API above UI).

    Each row carries a transient ``_section`` ("API"/"UI"). Rows are inserted (not
    appended at the sheet bottom) so API bugs stay in the API block and UI bugs in
    the UI block; the Status dropdown + colour rules span a fixed range (H2:H500),
    so inserted rows inherit them automatically.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path)  # not data_only → preserve styles
    ws = wb[_SHEET]
    hdr = {str(c.value).strip(): i for i, c in enumerate(ws[1], 1) if c.value}
    written = 0
    for section in _SECTIONS:
        batch = [r for r in rows if r.get("_section") == section]
        for row in batch:
            secs = _find_section_rows(ws)  # recompute — earlier inserts shift rows
            if section in secs:
                end = secs["UI"] if section == "API" and "UI" in secs else ws.max_row + 1
                idx = _last_bug_row(ws, secs[section], end) + 1
                ws.insert_rows(idx)
            else:  # fallback: no section header found → append at the bottom
                idx = ws.max_row + 1
            _write_row(ws, idx, row, hdr)
            written += 1
    # Any rows without a recognised section (shouldn't happen) → bottom-append.
    for row in (r for r in rows if r.get("_section") not in _SECTIONS):
        _write_row(ws, ws.max_row + 1, row, hdr)
        written += 1
    wb.save(path)
    return written


def reconcile(
    groups,
    *,
    passed_tc_ids: set[int] | None = None,
    tracker_path: Path = DEFAULT_TRACKER,
    allow_write: bool | None = None,
):
    """Compare app_bug failures to the tracker; append new rows on local runs.

    ``passed_tc_ids`` (the run's PASSED test-case ids) drives the RESOLVED report:
    an open bug whose TC passed this run is no longer reproducing → suggest closing
    it. Detection is suggest-only — the tracker's Status is never auto-changed.
    """
    res = Reconciliation(tracker=tracker_path)
    existing = load_tracker(tracker_path)
    is_ci = bool(os.environ.get("CI"))
    if allow_write is None:
        allow_write = (not is_ci) and tracker_path.exists()

    seen: set[int] = set()
    candidates: list[tuple[int, object]] = []
    for g in groups:
        if getattr(g, "category", "") != "app_bug":
            continue  # only real defects (API + UI) — skip env/flaky/deploy noise
        for tc in getattr(g, "tcs", []):
            n = _tc_int(tc)
            if n is not None and n not in seen:
                seen.add(n)
                candidates.append((n, g))

    reg = _registry_lookup()
    # Per-section running sequence (API / UI numbered independently).
    next_seq = {s: _next_seq(existing, s) for s in _SECTIONS}
    to_append: list[dict] = []

    def _build_entry(tc_id: int, g: object) -> dict:
        tc = reg.get(tc_id)  # enrich from the code registry when available
        name = getattr(tc, "tc_string", "") if tc else ""
        section = _section_for(tc_id, name)
        next_seq[section] += 1
        bug_id = f"BUG-{section}-{next_seq[section]:03d}"
        title = ((getattr(tc, "title", "") if tc else "") or "").strip()
        evidence = (getattr(g, "evidence", "") or "")[:300]
        expected, actual = _expected_actual(evidence)
        # Expected must never be blank. If the assertion had no "Expected … got …"
        # pattern (e.g. UI defects: overflow, "accepts invalid …"), fall back to the
        # TC title — it states the intended/correct behavior — then to the evidence.
        if not expected:
            expected = title or evidence[:200] or "(expected behavior — see Test Case Name)"
        return {
            "Bug ID": bug_id,
            "_section": section,  # transient hint for _append_rows (not a column)
            "Test Case ID": tc_id,
            "Test Case Name": name,
            "Module": getattr(tc, "module", "").capitalize() if tc else "",
            "Feature": _feature_from_tc_string(name),
            "Severity": "Major",
            "Priority": getattr(tc, "priority", "P2") if tc else "P2",
            "Status": "Open",
            "Reported By": "AI Triage",
            "Date Opened": datetime.date.today().isoformat(),
            "Summary": title or evidence[:200],
            "Expected": expected,
            "Actual": actual or evidence[:250],
            "Evidence / Assertion": evidence,
            "Notes": "",  # left blank by request — no auto-generated note text
        }

    for tc_id, g in sorted(candidates):
        row = existing.get(tc_id)
        cur_ev = (getattr(g, "evidence", "") or "")[:300]
        if row is None:
            entry = _build_entry(tc_id, g)
            res.new.append(entry)
            to_append.append(entry)
        elif _is_closed(row.status):
            # A closed bug is failing again. If the failure matches the recorded
            # evidence (same root cause) → REOPEN the same bug. If it's a DIFFERENT
            # failure (or the old evidence is unknown to compare) → treat SAME-cause
            # as reopen (conservative, no duplicate); only a clearly-different
            # signature opens a NEW bug distinct from the closed one.
            same_cause = (not row.evidence) or (_sig(cur_ev) == _sig(row.evidence))
            if same_cause:
                res.reopen.append(row)
            else:
                entry = _build_entry(tc_id, g)
                entry["_distinct_from"] = row.bug_id  # transient hint for render (not a column)
                res.new.append(entry)
                to_append.append(entry)
        else:
            res.known_open.append(row)

    # RESOLVED (mirror of REOPEN): an open bug whose TC PASSED this run no longer
    # reproduces → suggest closing it. Skip any TC that also failed this run (a
    # fail can't coexist with a pass, but guard anyway — a fail always wins).
    for tc_id in sorted(passed_tc_ids or ()):
        if tc_id in seen:
            continue
        row = existing.get(tc_id)
        if row is not None and not _is_closed(row.status):
            res.resolved.append(row)

    if to_append and allow_write:
        res.written = _append_rows(tracker_path, to_append)
        res.wrote_to_file = True
    return res


def render(res: Reconciliation) -> str:
    """Human-readable reconciliation block (safe for console + markdown).

    Sections are ordered most-actionable first (NEW → REOPEN → KNOWN OPEN →
    RESOLVED); RESOLVED is a close-suggestion, so it sits last. Within each section
    rows are sorted by Bug ID so the list reads in order (BUG-API-001, 002, …).
    """
    lines = ["", "── Bug Tracker reconciliation ──"]
    if not (res.new or res.known_open or res.reopen or res.resolved):
        lines.append("  No defects to track. ✅")
        return "\n".join(lines)
    for e in sorted(res.new, key=lambda e: e["Bug ID"]):
        tag = "added" if res.wrote_to_file else "would add (propose-only)"
        distinct = e.get("_distinct_from")
        extra = f"; NEW defect, different from closed {distinct}" if distinct else ""
        lines.append(f"  🆕 NEW         {e['Bug ID']}  (TC-{e['Test Case ID']}) — {tag}{extra}")
    for r in sorted(res.reopen, key=lambda r: r.bug_id):
        lines.append(
            f"  ⚠  REOPEN      {r.bug_id}  (TC-{r.tc_id}) — CLOSED but the SAME failure is "
            f"back → reopen"
        )
    for r in sorted(res.known_open, key=lambda r: r.bug_id):
        lines.append(f"  🔵 KNOWN OPEN  {r.bug_id}  (TC-{r.tc_id}) — already tracked, still open")
    for r in sorted(res.resolved, key=lambda r: r.bug_id):
        lines.append(
            f"  ✅ RESOLVED    {r.bug_id}  (TC-{r.tc_id}) — passed this run, no longer "
            f"failing → safe to CLOSE (status is '{r.status}')"
        )
    if res.new:
        if res.wrote_to_file:
            lines.append(f"  → {res.written} row(s) appended to {res.tracker.name}")
        else:
            lines.append(f"  → propose-only (run locally to write into {res.tracker.name})")
    return "\n".join(lines)
